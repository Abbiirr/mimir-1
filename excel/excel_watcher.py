# excel_watcher.py
import hashlib
import json
import logging
import threading
import time as pytime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple
import datetime as dt
import zipfile
from decimal import Decimal

from openpyxl import load_workbook
from watchdog.events import PatternMatchingEventHandler
from watchdog.observers import Observer  # swap to PollingObserver for network shares
from excel_kafka_publisher import publish_excel_changes, close_publisher

# -------- settings --------
WATCH_DIR = Path("./excels").resolve()
STATE_FILE = Path("./.excel_state.json")
DEBOUNCE_SECS = 0.8     # collapse bursts of FS events
SETTLE_SECS = 0.8       # wait between stats to ensure file isn't still writing
READ_TIMEOUT = 15       # max seconds to wait for settle
ALLOWED_EXTS = {".xlsx"}  # keep it simple for now
KEY_CANDIDATES = ("id", "ID", "Id")  # preferred stable row identifier headers

# -------- logging --------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s"
)
log = logging.getLogger("excel_watcher")

# -------- tiny state store --------
def _load_state() -> Dict[str, Dict[str, str]]:
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except Exception:
            log.warning("State file corrupted; starting fresh")
    return {}

def _save_state(state: Dict[str, Dict[str, str]]) -> None:
    STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

STATE: Dict[str, Dict[str, str]] = _load_state()
TIMERS: Dict[str, threading.Timer] = {}

# -------- helpers --------
def _jsonify_value(v):
    # Make values JSON-safe for logging (datetime -> ISO, Decimal -> float, bytes -> utf-8)
    if isinstance(v, (dt.datetime, dt.date, dt.time)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, bytes):
        return v.decode("utf-8", "replace")
    return v

def _jsonify_row(row: dict) -> dict:
    return {k: _jsonify_value(v) for k, v in row.items()}

def _normalize_for_hash(v):
    # Stable normalization for hashing keys/rows (strings only)
    if isinstance(v, (dt.datetime, dt.date, dt.time)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, bytes):
        return v.decode("utf-8", "replace")
    return "" if v is None else str(v)

def _is_temp(p: Path) -> bool:
    # Office lock/temp files (e.g., ~$Book.xlsx) and .tmp; also skip hidden dotfiles
    return p.name.startswith("~$") or p.suffix.lower() == ".tmp" or p.name.startswith(".")

def _wait_until_settled(p: Path, timeout: int = READ_TIMEOUT) -> bool:
    """Wait until file size/mtime stop changing for SETTLE_SECS; return False if vanished or timed out."""
    deadline = pytime.time() + timeout
    last: Tuple[int, float] = (-1, -1.0)
    while pytime.time() < deadline:
        try:
            st = p.stat()
            cur = (st.st_size, st.st_mtime)
        except FileNotFoundError:
            return False
        if cur == last:
            return True
        last = cur
        pytime.sleep(SETTLE_SECS)
    return False

def _sha1(values: Iterable) -> str:
    # Already stringifies values (safe for datetimes)
    b = json.dumps([None if v is None else str(v) for v in values], ensure_ascii=False).encode()
    return hashlib.sha1(b).hexdigest()

def _row_key_from(header: List[str], values: List[object]) -> str:
    row = {header[i]: values[i] if i < len(values) else None for i in range(len(header))}
    # Prefer natural key column if present
    for k in KEY_CANDIDATES:
        v = row.get(k)
        if v not in (None, "", " "):
            return str(v)
    # Fallback: stable hash of whole row (normalize first to avoid JSON datetime errors)
    norm_row = {k: _normalize_for_hash(row.get(k)) for k in header}
    return hashlib.md5(json.dumps(norm_row, sort_keys=True, ensure_ascii=False).encode()).hexdigest()

def _to_rows(path: Path) -> List[Tuple[str, str, str, Dict[str, object]]]:
    """
    Returns list of (sheet_key, row_hash, sheet_name, row_dict)
    sheet_key is namespaced like "Sheet1:<row_key>"
    """
    # Guard: ensure real OOXML .xlsx (ZIP container) to avoid BadZipFile
    if path.suffix.lower() == ".xlsx" and not zipfile.is_zipfile(path):
        log.error("[%s] Not a valid .xlsx (Open XML ZIP). Save as Excel Workbook (.xlsx).", path.name)
        return []

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)  # streaming + computed values
    out: List[Tuple[str, str, str, Dict[str, object]]] = []
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header = None

            # first non-empty row as header
            for r in rows:
                if r and any(c not in (None, "", " ") for c in r):
                    header = [str(c).strip() if c is not None else "" for c in r]
                    break
            if not header:
                continue

            for r in rows:
                if r is None:
                    continue
                vals = list(r) + [None] * (len(header) - len(r))
                row = {header[i]: vals[i] for i in range(len(header))}
                # skip fully-empty rows
                if not any(v not in (None, "", " ") for v in row.values()):
                    continue
                rk = _row_key_from(header, vals)
                rh = _sha1(vals[:len(header)])
                out.append((f"{ws.title}:{rk}", rh, ws.title, row))
    finally:
        wb.close()
    return out

def _diff_and_log(path: Path) -> None:
    if not path.exists() or _is_temp(path) or path.suffix.lower() not in ALLOWED_EXTS:
        return
    if not _wait_until_settled(path):
        log.warning("Unsettled save; skipping this round: %s", path.name)
        return

    try:
        rows = _to_rows(path)
    except Exception as e:
        log.error("Failed to read %s: %s", path.name, e)
        return

    prev = STATE.get(str(path), {})  # {sheet:key -> hash}
    now_map = {k: h for (k, h, _sheet, _row) in rows}

    new_rows, upd_rows = [], []
    for sheet_key, h, sheet, row in rows:
        old = prev.get(sheet_key)
        if old is None:
            new_rows.append((sheet, sheet_key, row))
        elif old != h:
            upd_rows.append((sheet, sheet_key, row))

    # print concise logs (JSON-safe)
    if new_rows:
        log.info("[%s] NEW rows: %d", path.name, len(new_rows))
        for sheet, sk, row in new_rows[:5]:
            log.info("  + %s %s %s", sheet, sk, json.dumps(_jsonify_row(row), ensure_ascii=False)[:200])
        if len(new_rows) > 5:
            log.info("  ... and %d more new rows", len(new_rows) - 5)

    if upd_rows:
        log.info("[%s] UPDATED rows: %d", path.name, len(upd_rows))
        for sheet, sk, row in upd_rows[:5]:
            log.info("  ~ %s %s %s", sheet, sk, json.dumps(_jsonify_row(row), ensure_ascii=False)[:200])
        if len(upd_rows) > 5:
            log.info("  ... and %d more updated rows", len(upd_rows) - 5)

    # persist new state
    STATE[str(path)] = now_map
    _save_state(STATE)
    publish_excel_changes(path, new_rows, upd_rows)

# -------- watchdog handler --------
class ExcelHandler(PatternMatchingEventHandler):
    def __init__(self):
        super().__init__(patterns=["*.xlsx"], ignore_patterns=["~$*"], ignore_directories=True)

    def _debounce(self, p: Path):
        s = str(p)
        t = TIMERS.get(s)
        if t:
            t.cancel()
        TIMERS[s] = threading.Timer(DEBOUNCE_SECS, lambda: _diff_and_log(p))
        TIMERS[s].start()

    def on_created(self, e):
        p = Path(e.src_path)
        if not _is_temp(p):
            log.info("created: %s", p.name)
            self._debounce(p)

    def on_modified(self, e):
        p = Path(e.src_path)
        if not _is_temp(p):
            log.info("modified: %s", p.name)
            self._debounce(p)

    def on_moved(self, e):
        # many editors save via temp + atomic rename => treat as fresh content
        p = Path(e.dest_path)
        if not _is_temp(p):
            log.info("moved-> %s", p.name)
            self._debounce(p)

# -------- bootstrap --------
def _seed_existing():
    for f in WATCH_DIR.glob("*.xlsx"):
        if not _is_temp(f):
            _diff_and_log(f)

def main():
    WATCH_DIR.mkdir(parents=True, exist_ok=True)
    log.info("Watching directory: %s", WATCH_DIR)
    _seed_existing()

    handler = ExcelHandler()
    obs = Observer()  # for flaky/network shares: from watchdog.observers.polling import PollingObserver; obs = PollingObserver()
    obs.schedule(handler, str(WATCH_DIR), recursive=False)
    obs.start()

    try:
        while True:
            pytime.sleep(1)
    except KeyboardInterrupt:
        log.info("Shutting down watcher...")
    finally:
        obs.stop()
        obs.join()
        close_publisher()

if __name__ == "__main__":
    main()
